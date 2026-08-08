"""Build anonymised all-school, all-subject score distributions from VCAA data."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


SOURCE_SUBJECT_BY_CODE: dict[str, str] = {
    "AC": "Accounting",
    "AH": "Agricultural and Horticultural Studies",
    "AL03": "Algorithmics (HESS)",
    "IT02": "Applied Computing: Data Analytics",
    "IT03": "Applied Computing: Software Development",
    "AT": "Art Creative Practice",
    "SA": "Art Making and Exhibiting",
    "BI": "Biology",
    "BM": "Business Management",
    "CH": "Chemistry",
    "CC": "Classical Studies",
    "DA": "Dance",
    "DR": "Drama",
    "EC": "Economics",
    "EN": "English",
    "EF": "English as an Additional Language",
    "EG": "English Language",
    "EV": "Environmental Science",
    "XI03": "Extended Investigation",
    "FT": "Food Studies",
    "GE": "Geography",
    "HH": "Health and Human Development",
    "HI17": "Ancient History",
    "HA": "Australian History",
    "HR": "History Revolutions",
    "LS": "Legal Studies",
    "LI": "Literature",
    "MA10": "Foundation Mathematics",
    "NF": "General Mathematics",
    "NJ": "Mathematical Methods",
    "NS": "Specialist Mathematics",
    "ME": "Media",
    "OS": "Outdoor and Environmental Studies",
    "PL": "Philosophy",
    "PE": "Physical Education",
    "PH": "Physics",
    "PS06": "Politics",
    "DT": "Product Design and Technologies",
    "PY": "Psychology",
    "RS": "Religion and Society",
    "SO03": "Sociology",
    "SE03": "Systems Engineering",
    "TS": "Theatre Studies",
    "VC": "Visual Communication Design",
    "AR": "Languages:Arabic",
    "CN": "Languages:Chinese First Language",
    "LO57": "Languages:Chinese Culture and Society",
    "CK": "Languages:Chinese Second Language Advanced",
    "CL": "Languages:Chinese Second Language",
    "FR": "Languages:French",
    "GN": "Languages:German",
    "MG": "Languages:Greek",
    "HI": "Languages:Hindi",
    "IX": "Languages:Indonesian Second Language",
    "IL": "Languages:Italian",
    "JS": "Languages:Japanese Second Language",
    "KS": "Languages:Korean Second Language",
    "LA": "Languages:Latin",
    "PN": "Languages:Persian",
    "LO49": "Languages:Punjabi",
    "RU": "Languages:Russian",
    "SP": "Languages:Spanish",
    "TU": "Languages:Turkish",
    "LO31": "Languages:Vietnamese Second Language",
}

MANUAL_SCHOOL_SOURCES: dict[str, str] = {
    "Haileybury College (Girls)": "Haileybury Girls College, Keysborough",
    "Camberwell Girls Grammar School": "Camberwell Anglican Girls Grammar School, Canterbury",
    "St Kevin's College": "St Kevin's College Toorak",
    "The King David School": "The King David School - Senior School, Armadale",
    "Yesodei HaTorah College": "Yesodei Hatorah College - Ormond Campus, Brighton",
    "Victorian College of the Arts": "Victorian College of the Arts Secondary School, Southbank",
    "Caulfield Grammar School,ST KILDA EAST": "Caulfield Grammar School - Caulfield Campus, St Kilda East",
    "Ivanhoe Grammar School,Doreen": "Ivanhoe Grammar School - Plenty Campus, Mernda",
    "Ivanhoe Grammar School,Ivanhoe": "Ivanhoe Grammar School",
}


def normalise_school_name(value: str, include_locality: bool = True) -> str:
    ascii_value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    comparable_value = ascii_value if include_locality else ascii_value.split(",")[0]
    return re.sub(r"[^a-z0-9]+", "", comparable_value.lower())


def split_school_locality(school_name: str) -> tuple[str, str]:
    school_parts = school_name.rsplit(",", maxsplit=1)
    if len(school_parts) == 1:
        return school_name, ""
    return school_parts[0].strip(), school_parts[1].strip()


def build_score_distributions(
    workbook_path: Path,
    schools_path: Path,
) -> tuple[dict[str, dict[str, list[int]]], dict[str, str], list[str]]:
    schools = json.loads(schools_path.read_text(encoding="utf-8"))
    worksheet = load_workbook(workbook_path, read_only=True, data_only=True).active
    source_rows = worksheet.iter_rows(values_only=True)
    scores_by_school_and_subject: dict[str, dict[str, list[int]]] = {}
    source_schools: set[str] = set()
    current_school_name: str | None = None

    for _, school_name, subject_name, study_score in source_rows:
        if school_name:
            current_school_name = school_name
            source_schools.add(school_name)
        if current_school_name is None or subject_name is None or study_score is None:
            continue
        school_scores = scores_by_school_and_subject.setdefault(current_school_name, {})
        school_scores.setdefault(subject_name, []).append(int(study_score))

    for school_scores in scores_by_school_and_subject.values():
        for subject_scores in school_scores.values():
            subject_scores.sort(reverse=True)

    source_schools_by_full_name = {
        normalise_school_name(school): school for school in source_schools
    }
    source_schools_by_base_name = {
        normalise_school_name(school, include_locality=False): school
        for school in source_schools
    }
    school_aliases: dict[str, str] = {}
    missing_schools: list[str] = []
    for local_school in schools:
        local_school_name = local_school["name"]
        source_school_name = MANUAL_SCHOOL_SOURCES.get(local_school_name)
        if source_school_name is None:
            source_school_name = source_schools_by_full_name.get(
                normalise_school_name(local_school_name)
            )
        if source_school_name is None:
            source_school_name = source_schools_by_base_name.get(
                normalise_school_name(local_school_name, include_locality=False)
            )
        if source_school_name is None:
            missing_schools.append(local_school_name)
            continue
        school_aliases[local_school_name] = source_school_name

    return (
        dict(sorted(scores_by_school_and_subject.items())),
        dict(sorted(school_aliases.items())),
        missing_schools,
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--schools", type=Path, default=Path("app/data/schools.json"))
    arguments = parser.parse_args()

    distributions, school_aliases, missing_schools = build_score_distributions(
        arguments.workbook,
        arguments.schools,
    )
    source = "https://www.vcaa.vic.edu.au/sites/default/files/2026-02/2025StudentData.xlsx"
    school_options = [
        {"name": school_name, "locality": split_school_locality(school_name)[1]}
        for school_name in distributions
    ]
    output = f'''/**
 * Every published 2025 VCE 40+ study score, grouped by school and subject.
 * Source: {source}
 *
 * Student names are intentionally excluded. Scores are sorted highest first and
 * form the published part of each school/subject rank curve.
 */
export type HonourRollSchoolOption = {{
  name: string;
  locality: string;
}};

export const HONOUR_ROLL_2025_SCHOOL_OPTIONS: readonly HonourRollSchoolOption[] = {json.dumps(school_options, ensure_ascii=False, indent=2)};

export const HONOUR_ROLL_2025_SCHOOL_ALIASES: Readonly<Record<string, string>> = {json.dumps(school_aliases, ensure_ascii=False, indent=2, sort_keys=True)};

export const HONOUR_ROLL_2025_SUBJECT_BY_CODE: Readonly<Record<string, string>> = {json.dumps(SOURCE_SUBJECT_BY_CODE, ensure_ascii=False, indent=2, sort_keys=True)};

export const HONOUR_ROLL_2025_SCORES: Readonly<
  Record<string, Readonly<Record<string, readonly number[]>>>
> = {json.dumps(distributions, ensure_ascii=False, indent=2, sort_keys=True)};

export function getHonourRollSchoolName(schoolName: string): string | null {{
  if (HONOUR_ROLL_2025_SCORES[schoolName]) {{
    return schoolName;
  }}
  return HONOUR_ROLL_2025_SCHOOL_ALIASES[schoolName] ?? null;
}}

export function getHonourRollStudyScores(
  schoolName: string,
  subjectCode: string,
): readonly number[] {{
  const sourceSchoolName = getHonourRollSchoolName(schoolName);
  const sourceSubjectName = HONOUR_ROLL_2025_SUBJECT_BY_CODE[subjectCode];
  if (sourceSchoolName === null || sourceSubjectName === undefined) {{
    return [];
  }}
  return HONOUR_ROLL_2025_SCORES[sourceSchoolName]?.[sourceSubjectName] ?? [];
}}
'''
    arguments.output.write_text(output, encoding="utf-8", newline="\n")
    print(f"Schools: {len(distributions)}")
    print(f"Published scores: {sum(len(scores) for subjects in distributions.values() for scores in subjects.values())}")
    print(f"Subjects: {len({subject for subjects in distributions.values() for subject in subjects})}")
    print("Schools without published anchors:", ", ".join(missing_schools))
    print("Mazenod General Mathematics:", distributions.get("Mazenod College, Mulgrave", {}).get("General Mathematics"))


if __name__ == "__main__":
    main()
